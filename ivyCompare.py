from openpyxl import load_workbook
from elo_compare import *
from tabulate import tabulate

filename = "ivy results/1V8 ivy results table.xlsx"

wb = load_workbook(filename)
ws = wb.active

colleges = ["Brown", "Columbia", "Cornell", "Dartmouth",
            "Harvard", "Penn", "Princeton", "Yale"]

elo = [1300 for i in colleges]
table = [colleges, elo]

for row in range(3,40):
    if ws['B' + str(row)].value == 'x':
        ratings = []
        times = []
        raceType = ws['C' + str(row)].value
        for col in range(5,13):
            ratings.append(elo[col - 5])
            if ws.cell(row = row, column = col).value != None:
                times.append(ws.cell(row = row, column = col).value*24*60)
            else:
                times.append(0)
        elo = eloAdjust(times, ratings, raceType)
        table.append(roundList(elo))
    for i, rating in enumerate(elo):
        ws.cell(row = row, column = i + 14, value = rating)

print tabulate(table)
wb.save(filename)
