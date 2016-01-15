from openpyxl import Workbook, load_workbook
from types import *
import datetime

wb = load_workbook('testseason.xlsx')
sheets = ['Wisconsin',
          'Brown',
          'Washington',
          'Stanford',
          'Pennsylvania',
          'Princeton',
          'BU',
          'Navy',
          'Columbia',
          'California',
          'Dartmouth',
          'Holy Cross',
          'Harvard',
          'FIT',
          'Northeastern',
          'Cornell',
          'G Washington',
          'Santa Clara',
          'OK City',
          'Oregon State',
          'Syracuse',
          'Drexel',
          'Yale',
          'Hobart'
          ]

def getTime(t):
    if type(t) == UnicodeType:
        try:
            return datetime.datetime.strptime(t[3:],'%M:%S.%f').time()
        except:
            return t
    else:
        return t

def getDate(d):
    try:
        return d.date()
    except:
        return d



races = {}
rs = wb["Races"]
for r in range(2,25):
    name = rs.cell(row=r, column=1).value
    date = rs.cell(row=r, column=3).value
    races[name] = {"teams":[],  # Teams racing
                   "date":date, # Date of race
                   "results":[] # Order of results 1V8,2V8,3V8,4V8,5V8,1F8,1V4
                   }
duals = []

for sheet in sheets:
    ws = wb[sheet]
    print sheet
    for col in range(3,21):
        race = ws.cell(row=2, column=col).value
        if not race:
            continue
        date = ws.cell(row=1, column=col).value
        res = []
        for i in range(3,10):
            res.append(ws.cell(row=i, column=col).value)
        if race.startswith("v "):
            notexists = True
            for dual in duals:
                if sheet in dual["teams"] and race[2:] in dual["teams"] and date == dual["date"]:
                    dual["results"].append(res)
                    notexists = False
            if notexists:
                duals.append({"teams":[sheet, race[2:]],
                              "date":date,
                              "results":[res]})
        else:
            races[race]["teams"].append(sheet)
            races[race]["results"].append(res)


print "\nRaces:"
for race in races:
    print race, races[race]["teams"]

print "\nDuals"
for dual in duals:
    print dual["teams"]
    for res in dual["results"]:
        if res[0]:
            print "V8:",res[0].strftime("%M:%S.%f")

