from openpyxl import load_workbook

def getAbbr(teamname):
    abbrs = {'Wisconsin':'Wisconsin',
              'Brown':'Brown',
              'Washington':'Washington',
              'Stanford':'Stanford',
              'Pennsylvania':'Penn',
              'Princeton':'Princeton',
              'BU':'BU',
              'Navy':'Navy',
              'Columbia':'Columbia',
              'California':'Cal',
              'Dartmouth':'Dartmouth',
              'Holy Cross':'HolyCross',
              'Harvard':'Harvard',
              'FIT':'FIT',
              'Northeastern':'NEastern',
              'Cornell':'Cornell',
              'G Washington':'GW',
              'Santa Clara':'SantaClara',
              'OK City':'OKCityU',
              'Oregon State':'OregonSt',
              'Syracuse':'Syracuse',
              'Drexel':'Drexel',
              'Yale':'Yale',
              'Hobart':'Hobart'
              }
    return abbrs[teamname]

def generateCMAX():
    wb = load_workbook("data/fullcmax.xlsx")
    ws = wb.active

    cmax = {}
    cmax["min"] = 0
    for col in range(3,26):
        date = ws.cell(row=3,column=col).value
        if date:
            cmax[date] = []
            for row in range(5,155):
                team = ws.cell(row=row, column=col).value
                if team:
                    rating = ws.cell(row=row, column=col+1).value
                    if rating != None:
                        cmax[date].append((team,rating))
            if cmax["min"] == 2:
                cmax["min"] = date
            elif type(cmax["min"]) == int:
                cmax["min"] += 1
    
    return cmax


def pointFormula(deltaCMAX, deltaMOV):
    if deltaCMAX >= 0:
        winner = deltaMOV+deltaCMAX + 10
        loser = deltaCMAX-deltaMOV - 2.5
        return (winner, loser)
    winner = deltaMOV+deltaCMAX + 7.5
    loser = deltaCMAX-deltaMOV + 5
    return (winner, loser)

