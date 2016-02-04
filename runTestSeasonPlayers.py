from openpyxl import Workbook, load_workbook
from elo_files.elo_compare import *
from Race import *
from calcPoints import generateCMAX
from getPlayers import getPlayers
import matplotlib.pyplot as plt
import datetime

sheets = ['Wisconsin',
          'Brown',
          'Washington',
          'Stanford',
          'Pennsylvania',
          'Princeton',
          'BU',
          'Navy',
          'Columbia',
          'California',
          'Dartmouth',
          'Holy Cross',
          'Harvard',
          'FIT',
          'Northeastern',
          'Cornell',
          'G Washington',
          'Santa Clara',
          'OK City',
          'Oregon State',
          'Syracuse',
          'Drexel',
          'Yale',
          'Hobart'
          ]

wb = load_workbook('testseason.xlsx')

def getRaces(sheets = ['Wisconsin',
              'Brown',
              'Washington',
              'Stanford',
              'Pennsylvania',
              'Princeton',
              'BU',
              'Navy',
              'Columbia',
              'California',
              'Dartmouth',
              'Holy Cross',
              'Harvard',
              'FIT',
              'Northeastern',
              'Cornell',
              'G Washington',
              'Santa Clara',
              'OK City',
              'Oregon State',
              'Syracuse',
              'Drexel',
              'Yale',
              'Hobart'
              ]):
    races = RaceList()
    duals = RaceList()
    rs = wb["Races"]
    for r in range(2,23):
        if r:
            name = rs.cell(row=r, column=1).value
            date = rs.cell(row=r, column=3).value
            races.addRace(Race(name, date))

    for sheet in sheets:
        ws = wb[sheet]
        for col in range(3,21):
            race = ws.cell(row=2, column=col).value
            if not race:
                continue
            date = ws.cell(row=1, column=col).value
            res = []
            for i in range(3,10):
                res.append(ws.cell(row=i, column=col).value)
            if race.startswith("v "):
                notexists = True
                for dual in duals:
                    if sheet in dual.teams and race[2:] in dual.teams and date == dual.date:
                        dual.addResult(sheet, res)
                        notexists = False
                if notexists:
                    duals.addRace(Race(sheet+" v "+race[2:],date,teams=[sheet, race[2:]],results={sheet:res}))
            else:
                races.addTeamTo(race, sheet)
                races.addResultTo(race, sheet, res)


    return (races, duals)

def main():
    races, duals = getRaces()
    players = getPlayers(wb, races, duals)

    ''' Print out race info and whatnot
    print "\nRaces:"
    for race in races:
        print race.name
        for team, results in race.results.iteritems():
            print team
            if results[0]:
                print results[0].strftime("%M:%S.%f")
        print ""

    print "\nDuals"
    for dual in duals:
        print dual.teams[0]+" v "+dual.teams[1]
        for team, res in dual.results.iteritems():
            if res[0]:
                print "V8:",res[0].strftime("%M:%S.%f")
        print ""

    '''

    cmax = generateCMAX()

    prompt = raw_input("?")
    while prompt != "q":
        p = players.getPlayer(prompt)
        print prompt
        if p:
            print p.getPointsDetailed(cmax)
        prompt = raw_input("?")


    wbout = Workbook()
    ws1 = wbout.active
    ws1.title = "Player Rankings"
    print len(players)

    outplayers = []
    for player in players:
        points = player.getPoints(cmax)
        outplayers.append([player.team, player.name, points, points/len(player.races)])

    for i, p in enumerate(outplayers):
        for j, e in enumerate(p):
            ws1.cell(row = i+1, column=j+1, value=e)
        print p
    wbout.save(filename = "test.xlsx")


if __name__=="__main__":
    main()
