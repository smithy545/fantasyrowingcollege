from openpyxl import load_workbook
from elo_compare import *
#from tabulate import tabulate

filename = "the big crews/1V8 big crew results table.xlsx"

wb = load_workbook(filename)
ws = wb.active

colleges = ["Brown", "Columbia", "Cornell", "Dartmouth",
            "Harvard", "Penn", "Princeton", "Yale",
            "Wash", "Cal", "Syracuse", "Navy",
            "BU", "Northeastern", "Stanford", "Wisco",
            "FIT", "Hobart", "Holy Cross", "G Washington",
            "Or State", "OK City", "Santa Clara", "Drexel"]

collegeELOs = {}
for college in colleges:
    collegeELOs[college] = [1300]

elo = [1300 for i in colleges]
table = [colleges, elo]
races = []

for row in range(2,45):
    if ws['A' + str(row)].value == 'x':
        ratings = []
        times = []
        raceType = ws['B' + str(row)].value
        races.append(ws['C' + str(row)].value)
        for col in range(4,25):
            ratings.append(elo[col - 4])
            if ws.cell(row = row, column = col).value != None:
                times.append(ws.cell(row = row, column = col).value*24*60)
            else:
                times.append(0)
        elo = eloAdjust(times, ratings, raceType)
        for i, rating in enumerate(elo):
            collegeELOs[colleges[i]].append(rating)
        table.append(roundList(elo))

    '''
    for i, rating in enumerate(elo):
        ws.cell(row = row, column = i + 29, value = rating)
    '''

for i, row in enumerate(table):
    if i > 1:
        row.insert(0, races[i - 2])
    else:
        row.insert(0, "")

#print tabulate(table)
wb.save(filename)
