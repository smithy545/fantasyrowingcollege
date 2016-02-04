from openpyxl import Workbook, load_workbook
from datetime import datetime

def getAbbr(teamname):
    abbrs = {'Wisconsin':'Wisconsin',
              'Brown':'Brown',
              'Washington':'Washington',
              'Stanford':'Stanford',
              'Pennsylvania':'Penn',
              'Princeton':'Princeton',
              'BU':'BU',
              'Navy':'Navy',
              'Columbia':'Columbia',
              'California':'Cal',
              'Dartmouth':'Dartmouth',
              'Holy Cross':'HolyCross',
              'Harvard':'Harvard',
              'FIT':'FIT',
              'Northeastern':'NEastern',
              'Cornell':'Cornell',
              'G Washington':'GW',
              'Santa Clara':'SantaClara',
              'OK City':'OKCityU',
              'Oregon State':'OregonSt',
              'Syracuse':'Syracuse',
              'Drexel':'Drexel',
              'Yale':'Yale',
              'Hobart':'Hobart'
              }
    return abbrs[teamname]

def generateCMAX():
    wb = load_workbook("fullcmax.xlsx")
    ws = wb.active

    cmax = {}

    for col in range(3,26):
        date = ws.cell(row=3,column=col).value
        if date:
            cmax[date] = []
            for row in range(5,155):
                team = ws.cell(row=row, column=col).value
                if team:
                    rating = ws.cell(row=row, column=col+1).value
                    if rating != None:
                        cmax[date].append((team,rating))
            cmax["max"] = date
    
    return cmax

def getcmax(team, date, cmaxTable):
    team = getAbbr(team)

    colDate = date
    for key in cmaxTable:
        if key != "max" and colDate < key:
            colDate = key

    if colDate == date:
        colDate = cmaxTable["max"]

    for t in cmaxTable[colDate]:
        if team == t[0]:
            return t[1]

    raise NameError("Could not find:"+team)


def pointFormula(deltaCMAX, deltaMOV):
    if deltaCMAX >= 0:
        winner = (deltaMOV-deltaCMAX)
        loser = (deltaCMAX-deltaMOV)
        return (winner, loser)
    winner = deltaMOV+deltaCMAX
    loser = deltaMOV+deltaCMAX
    return (winner, loser)
