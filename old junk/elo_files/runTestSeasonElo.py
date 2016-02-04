from openpyxl import Workbook, load_workbook
from elo_files.elo_compare import *
import matplotlib.pyplot as plt
import datetime

sheets = ['Wisconsin',
          'Brown',
          'Washington',
          'Stanford',
          'Pennsylvania',
          'Princeton',
          'BU',
          'Navy',
          'Columbia',
          'California',
          'Dartmouth',
          'Holy Cross',
          'Harvard',
          'FIT',
          'Northeastern',
          'Cornell',
          'G Washington',
          'Santa Clara',
          'OK City',
          'Oregon State',
          'Syracuse',
          'Drexel',
          'Yale',
          'Hobart'
          ]

def getRaces(sheets = ['Wisconsin',
              'Brown',
              'Washington',
              'Stanford',
              'Pennsylvania',
              'Princeton',
              'BU',
              'Navy',
              'Columbia',
              'California',
              'Dartmouth',
              'Holy Cross',
              'Harvard',
              'FIT',
              'Northeastern',
              'Cornell',
              'G Washington',
              'Santa Clara',
              'OK City',
              'Oregon State',
              'Syracuse',
              'Drexel',
              'Yale',
              'Hobart'
              ]):
    wb = load_workbook('testseason.xlsx')
    races = {}
    rs = wb["Races"]
    for r in range(2,25):
        name = rs.cell(row=r, column=1).value
        date = rs.cell(row=r, column=3).value
        races[name] = {"teams":[],  # Teams racing
                       "date":date, # Date of race
                       "results":[] # Order of results 1V8,2V8,3V8,4V8,5V8,1F8,1V4
                       }
    duals = []

    for sheet in sheets:
        ws = wb[sheet]
        for col in range(3,21):
            race = ws.cell(row=2, column=col).value
            if not race:
                continue
            date = ws.cell(row=1, column=col).value
            res = []
            for i in range(3,10):
                res.append(ws.cell(row=i, column=col).value)
            if race.startswith("v "):
                notexists = True
                for dual in duals:
                    if sheet in dual["teams"] and race[2:] in dual["teams"] and date == dual["date"]:
                        dual["results"].append(res)
                        notexists = False
                if notexists:
                    duals.append({"teams":[sheet, race[2:]],
                                  "date":date,
                                  "results":[res]})
            else:
                races[race]["teams"].append(sheet)
                races[race]["results"].append(res)

    duals.sort(key=lambda x:x['date'].strftime('%Y-%m-%d'))

    return (races, duals)

def main():
    races, duals = getRaces()

    print "\nRaces:"
    for race in races:
        print race, races[race]["teams"]

    print "\nDuals"
    for dual in duals:
        print dual["teams"]
        for res in dual["results"]:
            if res[0]:
                print "V8:",res[0].strftime("%M:%S.%f")

    elos = {}
    for sheet in sheets:
        elos[sheet] = [(900, datetime.datetime(2015,3,20))]

    tdelta = 0
    for dual in duals:
        if dual["results"][0][0]:
            res0 = datetime.datetime.combine(datetime.date.today(),dual["results"][0][0])
            res1 = datetime.datetime.combine(datetime.date.today(),dual["results"][1][0])
            team0 = dual["teams"][0]
            team1 = dual["teams"][1]
        if res0 and res1:
            if res0 > res1:
                tdelta = (res0-res1).total_seconds()
                elo0,elo1 = elos[team0][-1][0],elos[team1][-1][0]
                elo1,elo0 = eloCompare(elo1,elo0,tdelta)
                elos[team0].append((elo0,dual["date"]))
                elos[team1].append((elo1,dual["date"]))
                print team1 + u" over " + team0 + u" by " + unicode(tdelta) + u"s"
            else:
                tdelta = (res1-res0).total_seconds()
                elo0,elo1 = elos[team0][-1][0],elos[team1][-1][0]
                elo0,elo1 = eloCompare(elo0,elo1,tdelta)
                elos[team0].append((elo0,dual["date"]))
                elos[team1].append((elo1,dual["date"]))
                print team0 + u" over " + team1 + u" by " + unicode(tdelta) + u"s"

    fig, ax = plt.subplots()
    plt.ylabel("Elo Rating")
    plt.xlabel("Date of race")
    for team in elos:
        plt.xlim(datetime.datetime(2015,3,20),datetime.datetime(2015,6,1))
        x = [i[1] for i in elos[team]]
        ax.plot(x ,[xy[0] for xy in elos[team]], label=team, linewidth=3)

    legend = ax.legend(loc='right')
    plt.show()

if __name__=="__main__":
    main()
