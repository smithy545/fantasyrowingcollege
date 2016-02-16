from openpyxl import Workbook, load_workbook
from Races import getRaces
from Players import getPlayers
from util import *
import sqlite3

def main():
    sheets = ['Wisconsin',
              'Brown',
              'Washington',
              'Stanford',
              #'Pennsylvania',
              'Princeton',
              'BU',
              'Navy',
              'Columbia',
              'California',
              'Dartmouth',
              'Holy Cross',
              'Harvard',
              'FIT',
              'Northeastern',
              #'Cornell',
              'G Washington',
              #'Santa Clara',
              #'OK City',
              'Oregon State',
              'Syracuse',
              'Drexel',
              'Yale',
              'Hobart'
              ]

    wb = load_workbook('data/2015 Spring Season.xlsx')

    races, duals = getRaces(wb)
    players = getPlayers(wb, races, duals, sheets)
    '''
    # Print out race info and whatnot
    print "\nRaces:"
    for race in races:
        print race.name
        for team, results in race.results.iteritems():
            print team
            if results[0]:
                print results[0].strftime("%M:%S.%f")
        print ""

    # Print out dual info and whatnot
    print "\nDuals"
    for dual in duals:
        print dual.teams[0]+" v "+dual.teams[1]
        for team, res in dual.results.iteritems():
            if res[0]:
                print "V8:",res[0].strftime("%M:%S.%f")
        print ""
    '''

    # Generate player rankings and output them to "Player Rankings-generated.xlsx"
    cmax = generateCMAX()

    wbout = Workbook()
    ws1 = wbout.active
    ws1.title = "Player Rankings"

    outplayers = []
    for player in players:
        points = player.getPoints(cmax)
        first = player.name.split(' ')[0]
        last = player.name.split(' ')[1]
        outplayers.append([player.team, first, last, points, points/player.getNumCompetitors()])
        
        ps = wbout.create_sheet()          # Player sheet
        ps.title = player.name
        ps.cell(row=1,column=1,value="Name:")
        ps.cell(row=1,column=2,value=player.name)
        ps.cell(row=2,column=1,value="Team:")
        ps.cell(row=2,column=2,value=player.team)
        ps.cell(row=3,column=1,value="Avg Season Points:")
        ps.cell(row=3,column=2,value=points/len(player.races))
        ps.cell(row=4,column=1,value="Total Season Points:")
        ps.cell(row=4,column=2,value=points)

        schema = player.getSchema(cmax)
        curRow = 6
        for date, info in schema.iteritems():
            ps.cell(row=curRow,column=1,value=date)
            ps.cell(row=curRow,column=2,value="CMax")
            ps.cell(row=curRow,column=3,value="Time")
            pointsearned = 0
            curRow += 1
            for s in info:
                ps.cell(row=curRow,column=1,value=s["team"])
                ps.cell(row=curRow,column=2,value=s["cmax"])
                ps.cell(row=curRow,column=3,value=s["time"])
                pointsearned += s["points"]
                curRow += 1
            ps.cell(row=curRow,column=1,value="Points earned")
            ps.cell(row=curRow,column=2,value=pointsearned)
            curRow += 1
        

    ws1 = wbout["Player Rankings"]
    conn = sqlite3.connect("C:\\Users\\Philip Smith\\Documents\\github\\djangopractice\\rowing\\_notused\\teams.db")
    c = conn.cursor()
    
    for i, p in enumerate(outplayers):
        found = False
        stmt = u'SELECT * FROM athlete WHERE first_name="'+unicode(p[1])+u'" AND last_name="'+unicode(p[2])+u'"'
        #if c.execute(stmt).fetchone() == None:
        #   print p[0], p[1], p[2],"not found"
        for j, e in enumerate(p):
            ws1.cell(row = i+1, column=j+1, value=e)
            
    wbout.save(filename = "data/Player Rankings-generated.xlsx")
    conn.close()

if __name__=="__main__":
    main()

